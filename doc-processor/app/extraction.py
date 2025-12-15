from __future__ import annotations

import io
import logging
import re
import subprocess
import tempfile
from dataclasses import dataclass

import fitz  # PyMuPDF
from lxml import etree
from lxml import html as lxml_html

logger = logging.getLogger("processor.extraction")


@dataclass(frozen=True)
class ExtractedDocument:
    content_type: str | None
    pages_text: list[str]


def _xml_to_text(raw: bytes) -> str:
    # Best-effort: parse as XML and extract text nodes.
    parser = etree.XMLParser(recover=True, huge_tree=True)
    root = etree.fromstring(raw, parser=parser)
    text = " ".join(t.strip() for t in root.itertext() if t and t.strip())
    return text.strip()


def _html_to_text(raw: bytes) -> str:
    """
    Best-effort: parse HTML/XHTML and extract visible text.
    - Drops <script>/<style>/<noscript>
    - Collapses whitespace
    """
    try:
        doc = lxml_html.document_fromstring(raw)
    except Exception:
        # Fallback: try generic HTML parser
        parser = etree.HTMLParser(recover=True, huge_tree=True)
        root = etree.fromstring(raw, parser=parser)
        text = " ".join(t.strip() for t in root.itertext() if t and t.strip())
        return " ".join(text.split()).strip()

    # Remove non-text content
    for bad in doc.xpath("//script|//style|//noscript"):
        try:
            bad.drop_tree()
        except Exception:
            # best-effort
            pass

    # Preserve tables by converting them to Markdown blocks before flattening.
    return _html_to_markdown_with_tables(doc)


def _html_to_markdown_with_tables(doc: lxml_html.HtmlElement) -> str:
    """
    Convert HTML to a Markdown-ish text while preserving <table> structure.
    This is intentionally best-effort and dependency-free (no markdownify).
    """
    block_tags = {
        "p",
        "div",
        "section",
        "article",
        "header",
        "footer",
        "main",
        "aside",
        "nav",
        "h1",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "ul",
        "ol",
        "li",
        "pre",
        "blockquote",
        "hr",
    }

    def norm_ws(s: str) -> str:
        return " ".join((s or "").split()).strip()

    def render_table(table_el: etree._Element) -> str:
        rows = table_el.xpath(".//tr")
        grid: list[list[str]] = []
        for tr in rows:
            cells = tr.xpath("./th|./td")
            row: list[str] = []
            for c in cells:
                txt = norm_ws(" ".join(t for t in c.itertext() if t and t.strip()))
                txt = txt.replace("|", r"\|")
                row.append(txt)
            if any(x for x in row):
                grid.append(row)
        if not grid:
            return ""

        # Normalize row widths
        width = max(len(r) for r in grid)
        grid = [r + [""] * (width - len(r)) for r in grid]

        header = grid[0]
        sep = ["---"] * width
        body = grid[1:] if len(grid) > 1 else []

        def row_md(r: list[str]) -> str:
            return "| " + " | ".join(r) + " |"

        lines = [row_md(header), row_md(sep)]
        lines.extend(row_md(r) for r in body)
        return "\n".join(lines).strip()

    parts: list[str] = []

    def walk(node: etree._Element) -> None:
        tag = (getattr(node, "tag", "") or "").lower()

        if tag == "table":
            md = render_table(node)
            if md:
                parts.append("\n\n" + md + "\n\n")
            return  # don't recurse into table

        if tag == "br":
            parts.append("\n")
            return

        # node.text
        if getattr(node, "text", None):
            t = norm_ws(node.text)
            if t:
                parts.append(t + " ")

        for ch in list(node):
            walk(ch)
            if getattr(ch, "tail", None):
                t = norm_ws(ch.tail)
                if t:
                    parts.append(t + " ")

        if tag in block_tags:
            parts.append("\n\n")

    body = doc.find("body")
    root = body if body is not None else doc
    walk(root)

    s = "".join(parts)
    # cleanup: collapse spaces, preserve paragraphs
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+\n", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    return s.strip()


def _convert_office_to_pdf(raw: bytes, ext: str) -> bytes:
    """
    Convert DOC/DOCX to PDF via LibreOffice (headless).
    Returns PDF bytes. Raises on conversion failure.
    """
    ext = ext.lower().lstrip(".")
    if ext not in {"doc", "docx"}:
        raise ValueError(f"unsupported_office_ext:{ext}")

    with tempfile.TemporaryDirectory(prefix="docproc_") as td:
        in_path = f"{td}/input.{ext}"
        out_dir = td
        with open(in_path, "wb") as f:
            f.write(raw)

        # LibreOffice output file name is based on input basename.
        # Use --nologo/--nolockcheck for containers.
        cmd = [
            "soffice",
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            out_dir,
            in_path,
        ]
        p = subprocess.run(cmd, capture_output=True, text=True)
        if p.returncode != 0:
            raise RuntimeError(f"libreoffice_failed: rc={p.returncode} stderr={p.stderr[:500]}")

        pdf_path = f"{td}/input.pdf"
        with open(pdf_path, "rb") as f:
            return f.read()


def pdf_to_page_pngs(pdf_bytes: bytes, *, max_pages: int, max_side_px: int) -> list[bytes]:
    """
    Render PDF pages to PNG bytes using PyMuPDF.
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    out: list[bytes] = []
    n = min(doc.page_count, max_pages)
    for i in range(n):
        page = doc.load_page(i)
        # Scale so the largest side is max_side_px
        rect = page.rect
        max_side = max(rect.width, rect.height)
        zoom = max_side_px / max_side if max_side > 0 else 1.0
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        out.append(pix.tobytes("png"))
    return out


def extract_text_non_vlm(raw: bytes, content_type: str | None, filename: str | None) -> ExtractedDocument:
    """
    Fast extraction for XML/plain text without VLM.
    """
    ct = (content_type or "").split(";")[0].strip().lower() or None
    name = (filename or "").lower()

    # Excel spreadsheets (.xlsx)
    if ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or name.endswith(".xlsx"):
        try:
            return ExtractedDocument(content_type=ct or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", pages_text=[_xlsx_to_markdown(raw, filename)])
        except Exception as e:
            logger.warning("xlsx_parse_failed", extra={"extra": {"filename": filename, "error": str(e)}})
            # fall through to plain-ish decode

    if ct in {"application/xml", "text/xml"} or name.endswith(".xml"):
        return ExtractedDocument(content_type=ct, pages_text=[_xml_to_text(raw)])

    if ct in {"text/html", "application/xhtml+xml"} or name.endswith(".html") or name.endswith(".htm") or name.endswith(".xhtml"):
        return ExtractedDocument(content_type=ct or "text/html", pages_text=[_html_to_text(raw)])

    # Plain-ish text fallback
    try:
        text = raw.decode("utf-8", errors="replace").strip()
    except Exception:
        text = ""
    return ExtractedDocument(content_type=ct, pages_text=[text])


def _xlsx_to_markdown(raw: bytes, filename: str | None) -> str:
    """
    Best-effort XLSX -> Markdown tables.
    - Converts each sheet into a Markdown table.
    - Applies conservative limits to avoid exploding very large sheets.
    """
    # Import only when needed to keep startup lighter.
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl_not_installed") from e

    max_rows_per_sheet = 200
    max_cols_per_sheet = 50
    max_sheets = 10

    wb = load_workbook(filename=io.BytesIO(raw), data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)[:max_sheets]

    def cell_to_str(v: object) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            s = v.strip()
        else:
            s = str(v).strip()
        # Keep markdown table stable
        s = s.replace("\n", " ").replace("\r", " ")
        s = s.replace("|", r"\|")
        return s

    parts: list[str] = []
    if filename:
        parts.append(f"# {filename}\n")

    for si, sn in enumerate(sheet_names):
        ws = wb[sn]

        # Read rows (bounded)
        rows: list[list[str]] = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri >= max_rows_per_sheet:
                break
            rows.append([cell_to_str(v) for v in (list(row)[:max_cols_per_sheet])])

        # Trim empty trailing columns
        width = 0
        for r in rows:
            for j in range(len(r) - 1, -1, -1):
                if r[j]:
                    width = max(width, j + 1)
                    break
        if width <= 0:
            # empty sheet
            continue
        rows = [r[:width] for r in rows]

        # Trim empty trailing rows
        while rows and all(not c for c in rows[-1]):
            rows.pop()
        if not rows:
            continue

        # Header heuristic: use first row as header if it has any non-empty cell
        header = rows[0]
        if not any(header):
            header = [f"Col {i+1}" for i in range(width)]
            body = rows
        else:
            body = rows[1:]

        def row_md(r: list[str]) -> str:
            return "| " + " | ".join((c or "") for c in r) + " |"

        parts.append(f"## Sheet: {sn}\n")
        parts.append(row_md(header))
        parts.append(row_md(["---"] * width))
        for r in body:
            # skip completely empty rows
            if all(not c for c in r):
                continue
            parts.append(row_md(r))

        # separate sheets
        if si < len(sheet_names) - 1:
            parts.append("\n")

    text = "\n".join(parts).strip()
    if not text:
        return (filename or "xlsx") + ": (empty spreadsheet)"
    return text


def normalize_to_pdf(raw: bytes, content_type: str | None, filename: str | None) -> tuple[bytes | None, str | None]:
    """
    Convert inputs to PDF bytes when possible (pdf/doc/docx).
    Returns (pdf_bytes, normalized_content_type).
    """
    ct = (content_type or "").split(";")[0].strip().lower() or None
    name = (filename or "").lower()

    if ct == "application/pdf" or name.endswith(".pdf"):
        return (raw, "application/pdf")

    if ct in {
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    } or name.endswith(".docx") or name.endswith(".doc"):
        ext = "docx" if name.endswith(".docx") or ct == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" else "doc"
        return (_convert_office_to_pdf(raw, ext), "application/pdf")

    return (None, ct)





